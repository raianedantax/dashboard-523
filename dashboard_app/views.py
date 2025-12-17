from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from .models import Aluno, Boletim, Turma, Curso, Serie, Turno, Disciplina, AlunoTurma
from decimal import Decimal, InvalidOperation # Importação corrigida e garantida
import openpyxl
import xlrd
import re
import json
from django.http import HttpResponse
from django.db import transaction  # <-- A importação que faltava
import pandas as pd
import os
from django.conf import settings
import unicodedata # Biblioteca para lidar com normalização de texto (acentos)
from django.contrib.auth.decorators import login_required



def consulta_boletim(request):
    """
    Controla a lógica da página de consulta de boletim.
    Busca o aluno pela matrícula fornecida e organiza seus boletins por ano.
    """
    contexto = {
        'aluno': None,
        'boletins_por_ano': {},
        'erro': None
    }

    # Verifica se o formulário foi submetido (método POST)
    if request.method == 'POST':
        matricula = request.POST.get('matricula')
        if matricula:
            try:
                # Tenta encontrar o aluno no banco de dados
                aluno = Aluno.objects.get(matricula=matricula)
                contexto['aluno'] = aluno

                # Busca todos os boletins associados ao aluno
                boletins = Boletim.objects.filter(aluno_matricula=aluno).select_related(
                    'disciplina'
                ).order_by('-turma_ano', 'disciplina__descricao')

                # Organiza os boletins em um dicionário agrupado por ano
                for boletim in boletins:
                    ano = boletim.turma_ano
                    if ano not in contexto['boletins_por_ano']:
                        contexto['boletins_por_ano'][ano] = []
                    contexto['boletins_por_ano'][ano].append(boletim)

            except Aluno.DoesNotExist:
                # Se o aluno não for encontrado, define uma mensagem de erro
                contexto['erro'] = "Matrícula não encontrada. Por favor, verifique os dados e tente novamente."

    # Renderiza o template HTML, passando o contexto com os dados
    return render(request, 'dashboard_app/boletim_aluno.html', contexto)


def importacao_excel(request):
    #df = pd.read_excel('\media\planilhas\planilha_modelo.xlsx')
    df = pd.read_excel(os.path.join(settings.BASE_DIR, 'media', 'planilhas', 'planilha_modelo.xlsx'))

    #dados alunos
    df_dados_alunos= df.iloc[7:, 1:3].reset_index(drop=True)
    df_dados_alunos = df_dados_alunos.dropna()
    df_dados_alunos.columns = ["MATRÍCULA", "NOME"]

    #situação
    #df_situacao = df_dados.iloc[:, -1]
    df_situacao = df.iloc[7:, -1].reset_index(drop=True)
    df_situacao.name = "SITUAÇÃO"

    #materias
    materias = []
    series_materias = df.iloc[5, 3:-1].reset_index(drop=True)
    series_materias = series_materias.dropna()
    for materia in series_materias:
        materias.append(materia)
    
    #separando dados
    df_dados_materias = df.iloc[7:, 3:-1].reset_index(drop=True)
    colunas_por_materia = 11

    blocos = []
    contagem = 0
    for i in range(0, df_dados_materias.shape[1], colunas_por_materia):
       bloco = df_dados_materias.iloc[:, i:i+colunas_por_materia].copy()

       bloco.columns = ["B1", "B2", "R1", "B3", "B4", "R2", "MÉDIA ANUAL", 
                        "MÉDIA R FINAL", "MÉDIA FINAL", "FALTAS", "FALTAS %"]
       
       bloco["MATÉRIA"] = materias[contagem]
       contagem+=1
       bloco = pd.concat([df_dados_alunos, bloco, df_situacao], axis=1)
       blocos.append(bloco)
       
    #parte final
    bloco_final = pd.concat(blocos, axis=0)

    return(HttpResponse(bloco_final.to_html()))

# --- VIEW DE IMPORTAÇÃO ATUALIZADA ---

    if request.method != 'POST':
        return render(request, 'dashboard_app/importar_mapa.html')

    contexto = {}
    try:
        arquivo_excel = request.FILES['planilha']

        # --- VERIFICAÇÃO DE FORMATO ATUALIZADA ---
        # Garante que o ficheiro é .xlsx ou .xls antes de continuar.
        if not (arquivo_excel.name.endswith('.xlsx') or arquivo_excel.name.endswith('.xls')):
            raise ValueError("O formato do ficheiro é inválido. Por favor, envie uma planilha no formato .xlsx ou .xls.")
        
        # --- LÓGICA CONDICIONAL PARA LER A PLANILHA ---
        sheet = None
        # Se for um ficheiro .xlsx, usa openpyxl
        if arquivo_excel.name.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(arquivo_excel, data_only=True)
            sheet = workbook.active
        # Se for um ficheiro .xls, usa xlrd
        else:
            workbook = xlrd.open_workbook(file_contents=arquivo_excel.read())
            sheet = workbook.sheet_by_index(0)

        # 1. Extrair Curso
        curso_bruto = sheet['B3'].value if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet) else sheet.cell_value(2, 1)
        curso_bruto = curso_bruto or ''
        
        # 2. Extrair Ano
        ano_bruto = str(sheet['P4'].value if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet) else sheet.cell_value(3, 15))
        ano_bruto = ano_bruto or ''

        # 3. Extrair Série (LÓGICA CORRIGIDA)
        serie_bruta = str(sheet['P5'].value if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet) else sheet.cell_value(4, 15))
        serie_bruta = serie_bruta or ''
        
        # Usa regex para encontrar o primeiro dígito na string (ex: '1' de '1º Ano', '2' de '2º', '3' de '3.0')
        match_serie = re.search(r'\d', serie_bruta)
        serie_digito = match_serie.group(0) if match_serie else None
        
        # Converte o dígito para o formato esperado ("1º", "2º", "3º")
        serie_extraida = f"{serie_digito}º" if serie_digito else "Não encontrado"
        contexto['serie'] = serie_extraida

        # 4. Extrair Turno
        turno_bruto = (sheet['B4'].value if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet) else sheet.cell_value(3, 1))
        turno_bruto = (turno_bruto or '').upper()
        # --- O RESTO DA LÓGICA DE EXTRAÇÃO CONTINUA IGUAL ---
        match_curso = re.search(r'EM\s(.*?)\s\(', curso_bruto)
        curso_extraido = match_curso.group(1).strip() if match_curso else "Não encontrado"
        contexto['curso'] = curso_extraido

        ano_extraido = ano_bruto.split('.')[0]
        contexto['ano'] = ano_extraido

        serie_extraida = serie_bruta.split(' ')[0]
        contexto['serie'] = serie_extraida

        turno_extraido = "Não encontrado"
        if "MATUTINO" in turno_bruto:
            turno_extraido = "MATUTINO"
        elif "VESPERTINO" in turno_bruto:
            turno_extraido = "VESPERTINO"
        elif "NOTURNO" in turno_bruto:
            turno_extraido = "NOTURNO"
        contexto['turno'] = turno_extraido

        mapa_cod_curso = {
            "INFORMÁTICA": "5",
            "ELETROTÉCNICA": "4",
            "EDIFICAÇÕES": "2",
            "SEGURANÇA DO TRABALHO": "S"
        }
        mapa_cod_turno = {"MATUTINO": "1", "VESPERTINO": "2", "NOTURNO": "3"}
        mapa_cod_serie = {"1º": "1", "2º": "2", "3º": "3"}

        cod_curso = mapa_cod_curso.get(curso_extraido, "?")
        cod_turno = mapa_cod_turno.get(turno_extraido, "?")
        #cod_serie = mapa_cod_serie.get(serie_extraida, "?")
        cod_serie = serie_digito if serie_digito else "?"
        
        contexto['turma_id'] = f"{cod_curso}{cod_turno}{cod_serie}"

    except Exception as e:
        contexto['erro'] = str(e)

    return render(request, 'dashboard_app/importar_mapa.html', contexto)


    """Função auxiliar para obter o valor de uma célula, compatível com openpyxl e xlrd."""
    if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
        return sheet.cell(row=row + 1, column=col + 1).value
    else: # xlrd
        try:
            return sheet.cell_value(row, col)
        except IndexError:
            return None

def boletim_aluno(request):
    context = {}
    if request.method == 'POST':
        matricula = request.POST.get('matricula')
        try:
            aluno = Aluno.objects.get(matricula=matricula)
            boletins = Boletim.objects.filter(aluno_matricula=matricula).select_related('disciplina').order_by('-turma_ano', 'disciplina__descricao')
            
            boletins_por_ano = {}
            for boletim in boletins:
                ano = boletim.turma_ano
                if ano not in boletins_por_ano:
                    boletins_por_ano[ano] = []
                boletins_por_ano[ano].append(boletim)

            context['aluno'] = aluno
            context['boletins_por_ano'] = boletins_por_ano

        except Aluno.DoesNotExist:
            context['erro'] = f"Nenhum aluno encontrado com a matrícula '{matricula}'. Por favor, verifique o número digitado."
        except Exception as e:
            context['erro'] = f"Ocorreu um erro inesperado: {e}"
            
    return render(request, 'dashboard_app/boletim_aluno.html', context)



    context = {}
    if request.method == 'POST':
        matricula = request.POST.get('matricula')
        try:
            aluno = Aluno.objects.get(matricula=matricula)
            boletins = Boletim.objects.filter(aluno_matricula=matricula).select_related('disciplina').order_by('-turma_ano', 'disciplina__descricao')
            
            boletins_por_ano = {}
            for boletim in boletins:
                ano = boletim.turma_ano
                if ano not in boletins_por_ano:
                    boletins_por_ano[ano] = []
                boletins_por_ano[ano].append(boletim)

            context['aluno'] = aluno
            context['boletins_por_ano'] = boletins_por_ano

        except Aluno.DoesNotExist:
            context['erro'] = f"Nenhum aluno encontrado com a matrícula '{matricula}'. Por favor, verifique o número digitado."
        except Exception as e:
            context['erro'] = f"Ocorreu um erro inesperado: {e}"
            
    return render(request, 'dashboard_app/boletim_aluno.html', context)

    context = {}
    if request.method == 'POST':
        matricula = request.POST.get('matricula')
        try:
            aluno = Aluno.objects.get(matricula=matricula)
            boletins = Boletim.objects.filter(aluno_matricula=matricula).select_related('disciplina').order_by('-turma_ano', 'disciplina__descricao')
            
            boletins_por_ano = {}
            for boletim in boletins:
                ano = boletim.turma_ano
                if ano not in boletins_por_ano:
                    boletins_por_ano[ano] = []
                boletins_por_ano[ano].append(boletim)

            context['aluno'] = aluno
            context['boletins_por_ano'] = boletins_por_ano

        except Aluno.DoesNotExist:
            context['erro'] = f"Nenhum aluno encontrado com a matrícula '{matricula}'. Por favor, verifique o número digitado."
        except Exception as e:
            context['erro'] = f"Ocorreu um erro inesperado: {e}"
            
    return render(request, 'dashboard_app/boletim_aluno.html', context)

def _get_cell_value(sheet, row, col):
    """Função auxiliar para obter o valor de uma célula, compatível com openpyxl e xlrd."""
    if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
        return sheet.cell(row=row + 1, column=col + 1).value
    else: # xlrd
        try:
            return sheet.cell_value(row, col)
        except IndexError:
            return None

def _normalize_str(s):
    """
    Normaliza uma string: converte para maiúsculas, remove acentos e espaços extras.
    Isto torna a correspondência de nomes de disciplinas muito mais robusta.
    """
    if not s:
        return ""
    s = str(s).upper().strip()
    return "".join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )



    contexto = {}
    
    if 'dados_importacao' in request.session:
        try:
            dados_preview = json.loads(request.session['dados_importacao'])
            contexto.update(dados_preview['header'])
            contexto['alunos_dados'] = dados_preview['alunos']
        except (json.JSONDecodeError, KeyError):
            del request.session['dados_importacao']

    if request.method == 'POST':
        try:
            if 'confirmar_importacao' in request.POST:
                with transaction.atomic():
                    dados_importacao = json.loads(request.session.get('dados_importacao', '{}'))
                    periodo_importacao = request.POST.get('periodo_importacao')

                    if not dados_importacao or not periodo_importacao:
                        raise ValueError("Sessão expirada ou dados de importação não encontrados.")

                    mapeamento_periodo = {
                        'ate_b1': ['bimestre1', 'faltas', 'faltaspercent'],
                        'ate_b2': ['bimestre1', 'bimestre2', 'faltas', 'faltaspercent'],
                        'ate_r1': ['bimestre1', 'bimestre2', 'recusem1', 'faltas', 'faltaspercent'],
                        'ate_b3': ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'faltas', 'faltaspercent'],
                        'ate_b4': ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'bimestre4', 'faltas', 'faltaspercent'],
                        'ate_r2': ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'bimestre4', 'recusem2', 'faltas', 'faltaspercent'],
                        'final':  ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'bimestre4', 'recusem2', 'recfinal', 'final', 'faltas', 'faltaspercent']
                    }
                    campos_para_atualizar = mapeamento_periodo.get(periodo_importacao, [])
                    header = dados_importacao['header']
                    alunos_dados = dados_importacao['alunos']
                    
                    curso_obj = Curso.objects.get(descricao__icontains=header['curso'])
                    serie_obj = Serie.objects.get(descricao__startswith=header['serie'])
                    turno_obj = Turno.objects.get(descricao__iexact=header['turno'])
                    
                    turma_obj, criada_turma = Turma.objects.get_or_create(
                        id=header['turma_id'], ano=header['ano'], 
                        defaults={
                            'descricao': f"{header['turma_id']}",
                            'curso': curso_obj, 'serie': serie_obj, 'turno': turno_obj
                        }
                    )

                    # --- LÓGICA DE PROGRESSÃO REFINADA ---
                    # Executa sempre para turmas de 2º ou 3º ano.
                    if int(serie_obj.id) in [2, 3]:
                        ano_anterior = int(header['ano']) - 1
                        # O ID da turma anterior tem o mesmo código de curso e turno, mas a série é decrementada.
                        id_turma_anterior = f"{header['turma_id'][:-1]}{int(header['turma_id'][-1]) - 1}"
                        
                        turma_anterior = Turma.objects.filter(id=id_turma_anterior, ano=ano_anterior).first()
                        if turma_anterior:
                            turma_obj.turma_id = turma_anterior.id
                            turma_obj.turma_ano = turma_anterior.ano
                            turma_obj.save()
                    
                    disciplinas_db = { _normalize_str(d.descricao): d for d in Disciplina.objects.all() }

                    relatorio = {'alunos_criados': 0, 'alunos_atualizados': 0, 'boletins_atualizados': 0, 'disciplinas_nao_encontradas': set()}
                    for aluno_data in alunos_dados:
                        aluno_obj, criado_aluno = Aluno.objects.get_or_create(matricula=aluno_data['matricula'], defaults={'nome': aluno_data['nome']})
                        if criado_aluno: relatorio['alunos_criados'] += 1
                        else: relatorio['alunos_atualizados'] += 1

                        AlunoTurma.objects.get_or_create(aluno_matricula=aluno_obj, turma_id=turma_obj.id, turma_ano=turma_obj.ano)
                        
                        for boletim_data in aluno_data['boletins']:
                            nome_disciplina_planilha = boletim_data['disciplina']
                            nome_normalizado = _normalize_str(nome_disciplina_planilha)
                            disciplina_obj = disciplinas_db.get(nome_normalizado)

                            if disciplina_obj:
                                identificadores = {
                                    'aluno_matricula': aluno_obj,
                                    'disciplina': disciplina_obj,
                                    'turma_id': turma_obj.id,
                                    'turma_ano': turma_obj.ano
                                }
                                
                                valores_para_atualizar = {'status': aluno_data.get('situacao')}
                                campos_decimais = ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'bimestre4', 'recusem2', 'recfinal', 'final']
                                campos_inteiros = ['faltas', 'faltaspercent']
                                
                                for campo in campos_para_atualizar:
                                    valor = boletim_data.get(campo)
                                    if valor is not None and str(valor).strip() != '':
                                        try:
                                            if campo in campos_decimais:
                                                valor_str = str(valor).replace(',', '.').strip()
                                                valores_para_atualizar[campo] = Decimal(valor_str)
                                            elif campo in campos_inteiros:
                                                valor_str = str(valor).split(',')[0].split('.')[0].strip()
                                                valores_para_atualizar[campo] = int(valor_str)
                                        except (InvalidOperation, ValueError, TypeError):
                                            continue
                                
                                Boletim.objects.update_or_create(defaults=valores_para_atualizar, **identificadores)
                                relatorio['boletins_atualizados'] += 1
                            else:
                                relatorio['disciplinas_nao_encontradas'].add(nome_disciplina_planilha)
                    
                del request.session['dados_importacao']
                messages.success(request, "Importação concluída com sucesso!")

                if relatorio['disciplinas_nao_encontradas']:
                    nao_encontradas_lista = sorted(list(relatorio['disciplinas_nao_encontradas']))
                    messages.warning(request, f"Atenção: As seguintes disciplinas da planilha não foram encontradas no banco de dados e foram ignoradas: {', '.join(nao_encontradas_lista)}")
                    relatorio['disciplinas_nao_encontradas'] = nao_encontradas_lista

                contexto = {'relatorio': relatorio}

            elif 'planilha' in request.FILES:
                arquivo_excel = request.FILES['planilha']
                if not (arquivo_excel.name.endswith('.xlsx') or arquivo_excel.name.endswith('.xls')):
                    raise ValueError("O formato do ficheiro é inválido. Por favor, envie uma planilha no formato .xlsx ou .xls.")
                
                sheet = None
                if arquivo_excel.name.endswith('.xlsx'):
                    workbook = openpyxl.load_workbook(arquivo_excel, data_only=True)
                    sheet = workbook.active
                else:
                    workbook = xlrd.open_workbook(file_contents=arquivo_excel.read())
                    sheet = workbook.sheet_by_index(0)

                curso_bruto = _get_cell_value(sheet, 2, 1) or ''
                ano_entrada_bruto = str(_get_cell_value(sheet, 3, 15) or '')
                serie_bruta = str(_get_cell_value(sheet, 4, 15) or '')
                
                ano_entrada = int(ano_entrada_bruto.split('.')[0])
                match_serie = re.search(r'(\d)', serie_bruta)
                serie_digito = match_serie.group(1) if match_serie else '1'
                serie_int = int(serie_digito)
                ano_corrente = ano_entrada + (serie_int - 1)

                turno_bruto = (_get_cell_value(sheet, 3, 1) or '').upper()
                match_curso = re.search(r'EM\s(.*?)\s\(', curso_bruto)
                curso_extraido = match_curso.group(1).strip() if match_curso else "Não encontrado"
                
                serie_extraida = f"{serie_digito}º" if serie_digito else "Não encontrado"
                turno_extraido = "Não encontrado"
                if "MATUTINO" in turno_bruto: turno_extraido = "MATUTINO"
                elif "VESPERTINO" in turno_bruto: turno_extraido = "VESPERTINO"
                elif "NOTURNO" in turno_bruto: turno_extraido = "NOTURNO"
                
                contexto.update({'curso': curso_extraido, 'ano': ano_corrente, 'serie': serie_extraida, 'turno': turno_extraido})
                
                mapa_cod_curso = {"INFORMÁTICA": "5", "ELETROTÉCNICA": "4", "EDIFICAÇÕES": "2", "SEGURANÇA DO TRABALHO": "S"}
                mapa_cod_turno = {"MATUTINO": "1", "VESPERTINO": "2", "NOTURNO": "3"}
                cod_curso = mapa_cod_curso.get(curso_extraido.upper(), "?")
                cod_turno = mapa_cod_turno.get(turno_extraido, "?")
                cod_serie = serie_digito if serie_digito else "?"
                contexto['turma_id'] = f"{cod_curso}{cod_turno}{cod_serie}"
                
                # --- LÓGICA DE VERIFICAÇÃO DE PROGRESSÃO NA PRÉ-VISUALIZAÇÃO ---
                if serie_int in [2, 3]:
                    ano_anterior = ano_corrente - 1
                    id_turma_anterior = f"{cod_curso}{cod_turno}{serie_int - 1}"
                    turma_anterior_obj = Turma.objects.filter(id=id_turma_anterior, ano=ano_anterior).first()
                    if turma_anterior_obj:
                        contexto['turma_anterior_info'] = f"Encontrada: {turma_anterior_obj.descricao} ({turma_anterior_obj.ano})"
                    else:
                        contexto['turma_anterior_info'] = "Nenhuma turma correspondente encontrada para o ano anterior."
                else:
                    contexto['turma_anterior_info'] = "Turma de 1º ano, não requer progressão."

                disciplinas = []
                max_cols = sheet.ncols if hasattr(sheet, 'ncols') else sheet.max_column
                
                coluna_situacao = -1
                for col in range(max_cols):
                    cabecalho = _get_cell_value(sheet, 7, col)
                    if cabecalho and "SITUAÇÃO" in str(cabecalho).upper():
                        coluna_situacao = col
                        break
                col_final = coluna_situacao if coluna_situacao != -1 else max_cols
                
                col_atual = 3
                while col_atual < col_final:
                    nome_disciplina = _get_cell_value(sheet, 6, col_atual)
                    if nome_disciplina and str(nome_disciplina).strip():
                        nome_limpo = str(nome_disciplina).strip()
                        disciplinas.append({'nome': nome_limpo, 'col_inicio': col_atual})
                    col_atual += 11

                alunos_dados = []
                max_rows = sheet.nrows if hasattr(sheet, 'nrows') else sheet.max_row
                for row in range(8, max_rows):
                    matricula = _get_cell_value(sheet, row, 1)
                    nome = _get_cell_value(sheet, row, 2)
                    if not matricula or not str(matricula).strip() or not nome or not str(nome).strip(): break
                    aluno_atual = {'matricula': str(int(matricula)) if isinstance(matricula, (int, float)) else str(matricula), 'nome': str(nome).strip(), 'boletins': [], 'situacao': _get_cell_value(sheet, row, coluna_situacao) if coluna_situacao != -1 else "N/A"}
                    for disc in disciplinas:
                        col = disc['col_inicio']
                        aluno_atual['boletins'].append({'disciplina': disc['nome'], 'bimestre1': _get_cell_value(sheet, row, col), 'bimestre2': _get_cell_value(sheet, row, col + 1), 'recusem1': _get_cell_value(sheet, row, col + 2), 'bimestre3': _get_cell_value(sheet, row, col + 3), 'bimestre4': _get_cell_value(sheet, row, col + 4), 'recusem2': _get_cell_value(sheet, row, col + 5), 'recfinal': _get_cell_value(sheet, row, col + 7), 'final': _get_cell_value(sheet, row, col + 8), 'faltas': _get_cell_value(sheet, row, col + 9), 'faltaspercent': _get_cell_value(sheet, row, col + 10)})
                    alunos_dados.append(aluno_atual)
                contexto['alunos_dados'] = alunos_dados
                dados_para_salvar = {'header': {k: v for k, v in contexto.items() if k != 'alunos_dados'}, 'alunos': alunos_dados}
                request.session['dados_importacao'] = json.dumps(dados_para_salvar)

        except Exception as e:
            messages.error(request, f"Ocorreu um erro crítico durante o processamento: {e}")
    
    return render(request, 'dashboard_app/importar_mapa.html', contexto)

def importar_mapa_notas(request):
    contexto = {}
    
    if 'dados_importacao' in request.session:
        try:
            dados_preview = json.loads(request.session['dados_importacao'])
            contexto.update(dados_preview['header'])
            contexto['alunos_dados'] = dados_preview['alunos']
        except (json.JSONDecodeError, KeyError):
            del request.session['dados_importacao']

    if request.method == 'POST':
        try:
            if 'confirmar_importacao' in request.POST:
                with transaction.atomic():
                    dados_importacao = json.loads(request.session.get('dados_importacao', '{}'))
                    periodo_importacao = request.POST.get('periodo_importacao')

                    if not dados_importacao or not periodo_importacao:
                        raise ValueError("Sessão expirada ou dados de importação não encontrados.")

                    mapeamento_periodo = {
                        'ate_b1': ['bimestre1', 'faltas', 'faltaspercent'],
                        'ate_b2': ['bimestre1', 'bimestre2', 'faltas', 'faltaspercent'],
                        'ate_r1': ['bimestre1', 'bimestre2', 'recusem1', 'faltas', 'faltaspercent'],
                        'ate_b3': ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'faltas', 'faltaspercent'],
                        'ate_b4': ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'bimestre4', 'faltas', 'faltaspercent'],
                        'ate_r2': ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'bimestre4', 'recusem2', 'faltas', 'faltaspercent'],
                        'final':  ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'bimestre4', 'recusem2', 'recfinal', 'final', 'faltas', 'faltaspercent']
                    }
                    campos_para_atualizar = mapeamento_periodo.get(periodo_importacao, [])
                    header = dados_importacao['header']
                    alunos_dados = dados_importacao['alunos']
                    
                    curso_obj = Curso.objects.get(descricao__icontains=header['curso'])
                    serie_obj = Serie.objects.get(descricao__startswith=header['serie'])
                    turno_obj = Turno.objects.get(descricao__iexact=header['turno'])
                    
                    try:
                        turma_obj = Turma.objects.get(id=header['turma_id'], ano=header['ano'])
                        criada_turma = False
                    except Turma.DoesNotExist:
                        turma_obj = Turma.objects.create(
                            id=header['turma_id'], ano=header['ano'], 
                            descricao=f"{header['turma_id']}",
                            curso=curso_obj, serie=serie_obj, turno=turno_obj
                        )
                        criada_turma = True

                    if not criada_turma and int(serie_obj.id) in [2, 3]:
                        ano_anterior = int(header['ano']) - 1
                        id_turma_anterior = f"{header['turma_id'][0]}{header['turma_id'][1]}{int(header['turma_id'][2]) - 1}"
                        turma_anterior = Turma.objects.filter(id=id_turma_anterior, ano=ano_anterior).first()
                        if turma_anterior:
                            turma_obj.turma_id = turma_anterior.id
                            turma_obj.turma_ano = turma_anterior.ano
                            turma_obj.save()
                    
                    disciplinas_db = { _normalize_str(d.descricao): d for d in Disciplina.objects.all() }

                    relatorio = {'alunos_criados': 0, 'alunos_atualizados': 0, 'boletins_atualizados': 0, 'disciplinas_nao_encontradas': set()}
                    for aluno_data in alunos_dados:
                        aluno_obj, criado_aluno = Aluno.objects.get_or_create(matricula=aluno_data['matricula'], defaults={'nome': aluno_data['nome']})
                        if criado_aluno: relatorio['alunos_criados'] += 1
                        else: relatorio['alunos_atualizados'] += 1

                        AlunoTurma.objects.get_or_create(aluno_matricula=aluno_obj, turma_id=turma_obj.id, turma_ano=turma_obj.ano)
                        
                        for boletim_data in aluno_data['boletins']:
                            nome_disciplina_planilha = boletim_data['disciplina']
                            nome_normalizado = _normalize_str(nome_disciplina_planilha)
                            disciplina_obj = disciplinas_db.get(nome_normalizado)

                            if disciplina_obj:
                                # --- CORREÇÃO DEFINITIVA ---
                                # Substituímos update_or_create por uma lógica explícita
                                # para garantir que a chave composta completa seja usada.
                                
                                # 1. Define os campos que identificam o boletim de forma única
                                identificadores = {
                                    'aluno_matricula': aluno_obj,
                                    'disciplina': disciplina_obj,
                                    'turma_id': turma_obj.id,
                                    'turma_ano': turma_obj.ano
                                }
                                
                                # 2. Prepara um dicionário com os valores a serem atualizados
                                valores_para_atualizar = {'status': aluno_data.get('situacao')}
                                campos_decimais = ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'bimestre4', 'recusem2', 'recfinal', 'final']
                                campos_inteiros = ['faltas', 'faltaspercent']
                                
                                for campo in campos_para_atualizar:
                                    valor = boletim_data.get(campo)
                                    if valor is not None and str(valor).strip() != '':
                                        try:
                                            if campo in campos_decimais:
                                                valor_str = str(valor).replace(',', '.').strip()
                                                valores_para_atualizar[campo] = Decimal(valor_str)
                                            elif campo in campos_inteiros:
                                                valor_str = str(valor).split(',')[0].split('.')[0].strip()
                                                valores_para_atualizar[campo] = int(valor_str)
                                        except (InvalidOperation, ValueError, TypeError):
                                            continue
                                
                                # 3. Verifica se o boletim existe
                                boletim_obj = Boletim.objects.filter(**identificadores).first()
                                
                                if boletim_obj:
                                    # Se existe, FORÇA a atualização usando a chave composta completa no filtro
                                    Boletim.objects.filter(**identificadores).update(**valores_para_atualizar)
                                else:
                                    # Se não existe, cria um novo
                                    Boletim.objects.create(**identificadores, **valores_para_atualizar)
                                
                                relatorio['boletins_atualizados'] += 1
                            else:
                                relatorio['disciplinas_nao_encontradas'].add(nome_disciplina_planilha)
                    
                del request.session['dados_importacao']
                messages.success(request, "Importação concluída com sucesso!")

                if relatorio['disciplinas_nao_encontradas']:
                    nao_encontradas_lista = sorted(list(relatorio['disciplinas_nao_encontradas']))
                    messages.warning(request, f"Atenção: As seguintes disciplinas da planilha não foram encontradas no banco de dados e foram ignoradas: {', '.join(nao_encontradas_lista)}")
                    relatorio['disciplinas_nao_encontradas'] = nao_encontradas_lista

                contexto = {'relatorio': relatorio}

            elif 'planilha' in request.FILES:
                arquivo_excel = request.FILES['planilha']
                if not (arquivo_excel.name.endswith('.xlsx') or arquivo_excel.name.endswith('.xls')):
                    raise ValueError("O formato do ficheiro é inválido. Por favor, envie uma planilha no formato .xlsx ou .xls.")
                
                sheet = None
                if arquivo_excel.name.endswith('.xlsx'):
                    workbook = openpyxl.load_workbook(arquivo_excel, data_only=True)
                    sheet = workbook.active
                else:
                    workbook = xlrd.open_workbook(file_contents=arquivo_excel.read())
                    sheet = workbook.sheet_by_index(0)

                # --- Lógica de Extração (sem alterações) ---
                ano_entrada_bruto = str(_get_cell_value(sheet, 3, 15) or '')
                serie_bruta = str(_get_cell_value(sheet, 4, 15) or '')
                ano_entrada = int(ano_entrada_bruto.split('.')[0])
                match_serie = re.search(r'(\d)', serie_bruta)
                serie_digito = match_serie.group(1) if match_serie else '1'
                ano_corrente = ano_entrada + (int(serie_digito) - 1)
                
                curso_bruto = _get_cell_value(sheet, 2, 1) or ''
                turno_bruto = (_get_cell_value(sheet, 3, 1) or '').upper()
                match_curso = re.search(r'EM\s(.*?)\s\(', curso_bruto)
                curso_extraido = match_curso.group(1).strip() if match_curso else "Não encontrado"
                serie_extraida = f"{serie_digito}º"
                turno_extraido = "MATUTINO" if "MATUTINO" in turno_bruto else "VESPERTINO" if "VESPERTINO" in turno_bruto else "NOTURNO" if "NOTURNO" in turno_bruto else "Não encontrado"
                
                contexto.update({'curso': curso_extraido, 'ano': ano_corrente, 'serie': serie_extraida, 'turno': turno_extraido})
                
                mapa_cod_curso = {"INFORMÁTICA": "5", "ELETROTÉCNICA": "4", "EDIFICAÇÕES": "2", "SEGURANÇA DO TRABALHO": "S"}
                mapa_cod_turno = {"MATUTINO": "1", "VESPERTINO": "2", "NOTURNO": "3"}
                cod_curso = mapa_cod_curso.get(curso_extraido.upper(), "?")
                cod_turno = mapa_cod_turno.get(turno_extraido, "?")
                cod_serie = serie_digito
                contexto['turma_id'] = f"{cod_curso}{cod_turno}{cod_serie}"

                if int(serie_digito) in [2, 3]:
                    ano_anterior_prog = ano_corrente - 1
                    id_turma_anterior_prog = f"{cod_curso}{cod_turno}{int(cod_serie) - 1}"
                    turma_anterior_obj = Turma.objects.filter(id=id_turma_anterior_prog, ano=ano_anterior_prog).first()
                    contexto['turma_anterior_info'] = f"Encontrada turma correspondente: {turma_anterior_obj.descricao} ({turma_anterior_obj.ano})" if turma_anterior_obj else "Nenhuma turma anterior correspondente encontrada."

                disciplinas = []
                max_cols = sheet.ncols if hasattr(sheet, 'ncols') else sheet.max_column
                col_atual = 3
                while True:
                    nome_disciplina = _get_cell_value(sheet, 6, col_atual)
                    if nome_disciplina and str(nome_disciplina).strip():
                        disciplinas.append({'nome': str(nome_disciplina).strip(), 'col_inicio': col_atual})
                    else:
                        cabecalho_situacao = _get_cell_value(sheet, 7, col_atual)
                        if cabecalho_situacao and "SITUAÇÃO" in str(cabecalho_situacao).upper():
                            contexto['coluna_situacao'] = col_atual
                            break
                    if col_atual > max_cols: break
                    col_atual += 11
                
                alunos_dados = []
                max_rows = sheet.nrows if hasattr(sheet, 'nrows') else sheet.max_row
                for row in range(8, max_rows):
                    matricula = _get_cell_value(sheet, row, 1)
                    if not matricula or not str(matricula).strip(): break
                    nome = str(_get_cell_value(sheet, row, 2)).strip()
                    aluno_atual = {'matricula': str(int(matricula)) if isinstance(matricula, (int, float)) else str(matricula), 'nome': nome, 'boletins': []}
                    if 'coluna_situacao' in contexto:
                        aluno_atual['situacao'] = _get_cell_value(sheet, row, contexto['coluna_situacao'])
                    for disc in disciplinas:
                        col = disc['col_inicio']
                        aluno_atual['boletins'].append({'disciplina': disc['nome'], 'bimestre1': _get_cell_value(sheet, row, col), 'bimestre2': _get_cell_value(sheet, row, col + 1), 'recusem1': _get_cell_value(sheet, row, col + 2), 'bimestre3': _get_cell_value(sheet, row, col + 3), 'bimestre4': _get_cell_value(sheet, row, col + 4), 'recusem2': _get_cell_value(sheet, row, col + 5), 'recfinal': _get_cell_value(sheet, row, col + 7), 'final': _get_cell_value(sheet, row, col + 8), 'faltas': _get_cell_value(sheet, row, col + 9), 'faltaspercent': _get_cell_value(sheet, row, col + 10)})
                    alunos_dados.append(aluno_atual)
                contexto['alunos_dados'] = alunos_dados
                
                dados_para_salvar = {'header': {k: v for k, v in contexto.items() if k != 'alunos_dados'}, 'alunos': alunos_dados}
                request.session['dados_importacao'] = json.dumps(dados_para_salvar)

        except Exception as e:
            messages.error(request, f"Ocorreu um erro crítico durante o processamento: {e}")
    
    return render(request, 'dashboard_app/importar_mapa.html', contexto)





    """
    Página inicial do dashboard. Exibe uma lista de todos os cursos
    para iniciar a navegação.
    """
    cursos = Curso.objects.all().order_by('descricao')
    contexto = {
        'cursos': cursos,
    }
    return render(request, 'dashboard_app/dashboard.html', contexto)
